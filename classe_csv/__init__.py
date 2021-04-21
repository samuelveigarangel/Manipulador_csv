from openpyxl import Workbook
from openpyxl import load_workbook
import menu

class Xlsx:


    def __init__(self):
        self.arquivo_excel = Workbook()

    def criar_planilha(self):
            """
                Função para criar planilhas
            :return: retorna as planilhas e o nome do arquivo em .xlsx
            """
            nome_documento = str(input('Digite o nome do seu arquivo extensão .xlsx: '))
            nome_planilha = str(input('Digite o nome da planilha: (separe por virgula caso deseje inserir mais de uma planilha)\n')).strip().split(',')
            for x in nome_planilha:
                planilha = self.arquivo_excel.create_sheet(x)
            self.arquivo_excel.save(nome_documento+'.xlsx')
            return self.arquivo_excel, nome_documento


    def ver_planilha(self, arquivo):
            """
                Função para ver planilhas
            :param arquivo: parametro recebe o arquivo que contem as planilhas
            :return:
            """
            print(f'Suas planilhas são: {arquivo.sheetnames}')


    def carregar_planilha(self):
            """
                Função para carregar arquivo em .xlsx que contém planilhas
            :return: retorna o arquivo carregado e o nome do arquivo
            """
            #coloque o arquivo .xlsx na pasta dos arquivos do programa
            try:
                arq = input('Digite o nome do arquivo: ').strip()
            except TypeError:
                print('Erro. Digite o caminho corretamente')
            else:
                try:
                    pasta = 'C:/Users/samue/PycharmProjects/ManipularCSV/' + arq + '.xlsx'
                    arquivo_carg = load_workbook(pasta)
                except FileNotFoundError:
                    print('Arquivo não encontrado, tente novamente!')
                    return None, None
                else:
                    print('\033[0;32mArquivo carregado...\033[0;0m')
                    return arquivo_carg, arq


    def inserir_cabecalho(self, planilha, arq):
        """
            Função para inserir cabeçalho nas planilhas
        :param planilha: parametro recebe o arquivo que contem as planilhas
        :param arq: parametro que recebe o nome do arquivo
        :return:
        """
        self.ver_planilha(planilha)
        count = 1
        try:
            nome_planilha = str(input('Digite o nome da planilha que você deseja inserir o cabeçalho : '))
            #verifica se a planilha existe
            test = planilha[nome_planilha]
        except (KeyError):
            print('Planilha não encontrada. Verifique o nome da planilha novamente!')
        else:
            head = str(input(
                'Insira o nome do cabeçalho: (separe por virgula caso deseje inserir mais de um cabeçalho)\n')).strip().split(',')
            for x in head:
                planilha[nome_planilha].cell(row=1, column=count, value=x)
                count += 1
            try:
                planilha.save(arq+'.xlsx')
            except PermissionError:
                print('\033[031mArquivo aberto. Por favor, feche o arquivo para salvar.\033[0;0m')


    def inserir_dados(self, planilha, arq):
        """

        :param planilha: parametro recebe o arquivo que contem as planilhas
        :param arq: parametro que recebe o nome do arquivo
        :return:
        """

        self.ver_planilha(planilha)
        count_coluna = 1
        count_linha = 1
        try:
            nome_planilha = str(input('Digite o nome da planilha que você deseja inserir dados: '))
            #verifica planilha
            teste = planilha[nome_planilha]
        except KeyError:
            print('\033[031mPlanilha não encontrada. Verifique o nome da planilha novamente!\033[0;0m')
        else:
            for j in range(1, planilha[nome_planilha].max_column + 1):
                print(f'{planilha[nome_planilha].cell(row=1, column=j).value:<8}', end=" - ")
            print()
            while True:
                dados = str(input('Insira os dados: (separe por virgula caso deseje inserir mais de um cabeçalho)\n')).strip().split(',')
                for x in dados:
                    planilha[nome_planilha].cell(row=planilha[nome_planilha].max_row + count_linha, column=count_coluna, value=x)
                    count_coluna += 1
                    count_linha = 0
                try:
                    opção = input('Você deseja continuar a inserir dados? [S/N] ').lower().strip()[0]
                except:
                    print('\033[31mOpção inválida. Digite "s" para sim e "n" para não\033[0;0m')
                else:
                    if opção == 's':
                        count_linha += 1
                        count_coluna = 1
                        continue
                    elif opção == 'n':
                        break

            try:
                planilha.save(arq+'.xlsx')
            except PermissionError:
                print('\033[031mArquivo aberto. Por favor, feche o arquivo para salvar.\033[0;0m')

    def ver_conteudo(self, planilha):
        """
                Função para ver conteudo de planilhas
        :param planilha: parametro recebe o arquivo que contem as planilhas
        :return:
        """
        self.ver_planilha(planilha)
        try:
            nome_planilha = str(input('Digite o nome da planilha que você deseja inserir dados: '))
            #verifica planilha
            teste = planilha[nome_planilha]
        except KeyError:
            print('\033[031mPlanilha não encontrada. Verifique o nome da planilha novamente!\033[0;0m')
        else:
            try:
                max_linha = planilha[nome_planilha].max_row
                max_coluna = planilha[nome_planilha].max_column
                for i in range(1, max_linha + 1):
                    for j in range(1, max_coluna + 1):
                        print(f'{planilha[nome_planilha].cell(row=i, column=j).value:<8}', end=" - ")
                    print()
            except TypeError:
                print('\033[031mArquivo vazio.\033[0;0m')